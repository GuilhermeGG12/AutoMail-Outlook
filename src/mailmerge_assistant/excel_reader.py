from __future__ import annotations

from pathlib import Path
from zipfile import BadZipFile
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from mailmerge_assistant.config import REQUIRED_COLUMNS, WORKSHEET_NAME
from mailmerge_assistant.models import SpreadsheetRow, WorkbookData
from mailmerge_assistant.validators import normalize_header


class ExcelReaderError(ValueError):
    """Friendly Excel reader error."""


def missing_required_columns(headers: list[str]) -> list[str]:
    present = set(headers)
    return [column for column in REQUIRED_COLUMNS if column not in present]


def read_clientes_workbook(path: str | Path) -> WorkbookData:
    workbook_path = Path(path)
    if workbook_path.suffix.lower() != ".xlsx":
        raise ExcelReaderError("Selecione um arquivo Excel no formato .xlsx.")
    try:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    except FileNotFoundError as exc:
        raise ExcelReaderError(
            "Não foi possível encontrar a planilha selecionada. "
            "Se ela estiver no OneDrive, marque o arquivo como disponível neste dispositivo."
        ) from exc
    except PermissionError as exc:
        raise ExcelReaderError(
            "Não foi possível abrir a planilha selecionada porque o Windows negou acesso. "
            "Feche o arquivo no Excel e tente novamente."
        ) from exc
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ExcelReaderError(
            "Não foi possível abrir a planilha selecionada. "
            "Verifique se o arquivo é um .xlsx válido e se está baixado neste computador."
        ) from exc

    if WORKSHEET_NAME not in workbook.sheetnames:
        raise ExcelReaderError(
            'A planilha não possui uma aba chamada "Clientes". Verifique o arquivo selecionado.'
        )

    worksheet = workbook[WORKSHEET_NAME]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ExcelReaderError("A aba Clientes está vazia.")
    headers = [normalize_header(value) for value in header_row]
    missing = missing_required_columns(headers)
    if missing:
        raise ExcelReaderError("Colunas obrigatórias ausentes: " + ", ".join(missing))

    rows: list[SpreadsheetRow] = []
    for excel_row_number, row_values in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if _is_fully_empty(row_values):
            continue
        values: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if header:
                values[header] = row_values[index] if index < len(row_values) else None
        rows.append(SpreadsheetRow(row_number=excel_row_number, values=values))
    workbook.close()
    return WorkbookData(headers=headers, rows=rows)


def _is_fully_empty(values: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)
